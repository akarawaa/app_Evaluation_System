"""Evaluation score export to Excel (.xlsx) for offline analysis.

Two sheets, matching how HR/finance actually use this kind of export:
  "สรุป"      — one row per evaluation, pivot-table friendly (totals, %, status)
  "รายละเอียด" — one row per (evaluation, item), for drilling into a specific
                 criterion when a summary number needs investigating

Visibility is the same policy as list_all()/view_detail() — this reuses
_sees_all_evaluations/_can_view so an export never contains a row the caller
couldn't otherwise see one-by-one.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import CurrentUser
from app.services.evaluations import _actor_employee_id, _sees_all_evaluations

_SUMMARY_HEADERS = [
    "รหัสพนักงาน", "ชื่อ-นามสกุล", "ตำแหน่ง", "สาขา",
    "ผู้ประเมิน", "ชนิดการประเมิน", "สถานะ",
    "คะแนนประเมิน", "คะแนนเต็ม", "คะแนนการมา-ลา",
    "คะแนนรวม", "คิดเป็นร้อยละ",
    "วันที่สร้างใบ", "วันที่ปิดใบ",
]

_DETAIL_HEADERS = [
    "รหัสพนักงาน", "ชื่อ-นามสกุล", "หมวด", "หัวข้อ", "คะแนน",
]


def _autosize(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 4)


def _bold_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)


async def build_evaluations_excel(session: AsyncSession, user: CurrentUser) -> bytes:
    see_all = _sees_all_evaluations(user)
    actor_emp = None if see_all else await _actor_employee_id(session, user.id)

    rows = (await session.execute(text(
        "select ev.id, ev.kind, ev.status, ev.eval_score, ev.eval_max, "
        "ev.attendance_score, ev.total_score, ev.percentage, "
        "ev.created_at, ev.finalized_at, "
        "emp.emp_code, emp.full_name, emp.position, br.name as branch_name, "
        "evaluator.full_name as evaluator_name "
        "from evaluations ev "
        "join employees emp on emp.id = ev.employee_id "
        "left join branches br on br.id = emp.branch_id "
        "left join employees evaluator on evaluator.id = ev.evaluator_id "
        "where :see_all "
        "   or emp.id = :actor_emp "
        "   or emp.supervisor_id = :actor_emp "
        "   or emp.manager_id = :actor_emp "
        "order by ev.created_at desc"
    ), {"see_all": see_all, "actor_emp": actor_emp})).mappings().all()

    eval_ids = [str(r["id"]) for r in rows]
    items_by_eval: dict[str, list] = {eid: [] for eid in eval_ids}
    if eval_ids:
        item_rows = (await session.execute(text(
            "select i.evaluation_id, i.category_name, i.item_name, s.score "
            "from evaluation_items i "
            "left join evaluation_scores s on s.evaluation_item_id = i.id "
            "where i.evaluation_id = any(:ids) "
            "order by i.evaluation_id, i.category_order, i.item_order"
        ), {"ids": eval_ids})).mappings().all()
        for r in item_rows:
            items_by_eval[str(r["evaluation_id"])].append(r)

    wb = Workbook()
    summary = wb.active
    summary.title = "สรุป"
    _bold_header(summary, _SUMMARY_HEADERS)
    for r in rows:
        summary.append([
            r["emp_code"], r["full_name"], r["position"], r["branch_name"],
            r["evaluator_name"], r["kind"], r["status"],
            float(r["eval_score"]) if r["eval_score"] is not None else None,
            float(r["eval_max"]) if r["eval_max"] is not None else None,
            float(r["attendance_score"]) if r["attendance_score"] is not None else None,
            float(r["total_score"]) if r["total_score"] is not None else None,
            float(r["percentage"]) if r["percentage"] is not None else None,
            r["created_at"].strftime("%Y-%m-%d") if r["created_at"] else None,
            r["finalized_at"].strftime("%Y-%m-%d") if r["finalized_at"] else None,
        ])
    _autosize(summary, _SUMMARY_HEADERS)
    summary.freeze_panes = "A2"

    detail = wb.create_sheet("รายละเอียด")
    _bold_header(detail, _DETAIL_HEADERS)
    for r in rows:
        for it in items_by_eval[str(r["id"])]:
            detail.append([
                r["emp_code"], r["full_name"], it["category_name"], it["item_name"],
                float(it["score"]) if it["score"] is not None else None,
            ])
    _autosize(detail, _DETAIL_HEADERS)
    detail.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
