"""Evaluation score export to Excel — visibility scoping + shape of the file."""
from io import BytesIO

from openpyxl import load_workbook

from conftest import auth
from test_evaluation_lifecycle import _new, org  # noqa: F401


async def test_export_visible_to_subject_chain_hr(api, org):
    r = await api.post("/api/evaluations", headers=auth(org["sup"]), json=_new(org))
    eid = r.json()["id"]
    for it in r.json()["items"]:
        await api.put(f"/api/evaluations/{eid}/scores", headers=auth(org["sup"]),
                       json={"scores": [{"evaluation_item_id": it["id"], "score": 4}], "comments": []})

    for tok in (org["sup"], org["dept"], org["hr"], org["md"], org["emp"]):
        r = await api.get("/api/evaluations/export", headers=auth(tok))
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        wb = load_workbook(BytesIO(r.content))
        assert wb.sheetnames == ["สรุป", "รายละเอียด"]
        summary = wb["สรุป"]
        codes = [row[0].value for row in summary.iter_rows(min_row=2)]
        assert "E1" in codes


async def test_export_excludes_unrelated_evaluations(api, org):
    r = await api.post("/api/evaluations", headers=auth(org["sup"]), json=_new(org))
    assert r.status_code == 201, r.text

    r2 = await api.get("/api/evaluations/export", headers=auth(org["emp"]))
    wb = load_workbook(BytesIO(r2.content))
    summary = wb["สรุป"]
    codes = [row[0].value for row in summary.iter_rows(min_row=2)]
    assert codes == ["E1"]  # subject sees only their own row, not other tenants' data


async def test_export_detail_sheet_has_item_rows(api, org):
    r = await api.post("/api/evaluations", headers=auth(org["sup"]), json=_new(org))
    ev = r.json()
    eid = ev["id"]
    for it in ev["items"]:
        await api.put(f"/api/evaluations/{eid}/scores", headers=auth(org["sup"]),
                       json={"scores": [{"evaluation_item_id": it["id"], "score": 3.5}], "comments": []})

    r2 = await api.get("/api/evaluations/export", headers=auth(org["hr"]))
    wb = load_workbook(BytesIO(r2.content))
    detail = wb["รายละเอียด"]
    rows = list(detail.iter_rows(min_row=2, values_only=True))
    matching = [row for row in rows if row[0] == "E1"]
    assert len(matching) == 28  # operational template item count
    assert all(row[4] == 3.5 for row in matching)


async def test_export_status_filter(api, org):
    r = await api.post("/api/evaluations", headers=auth(org["sup"]), json=_new(org))
    eid = r.json()["id"]

    r_finalized = await api.get("/api/evaluations/export", headers=auth(org["hr"]), params={"status": "finalized"})
    wb = load_workbook(BytesIO(r_finalized.content))
    assert wb["สรุป"].max_row == 1  # header only, this draft doesn't match

    r_draft = await api.get("/api/evaluations/export", headers=auth(org["hr"]), params={"status": "draft"})
    wb2 = load_workbook(BytesIO(r_draft.content))
    codes = [row[0].value for row in wb2["สรุป"].iter_rows(min_row=2)]
    assert "E1" in codes


async def test_export_date_range_filter(api, org):
    await api.post("/api/evaluations", headers=auth(org["sup"]), json=_new(org))

    r_future = await api.get("/api/evaluations/export", headers=auth(org["hr"]),
                             params={"date_from": "2999-01-01"})
    wb = load_workbook(BytesIO(r_future.content))
    assert wb["สรุป"].max_row == 1  # nothing created that far in the future

    r_past = await api.get("/api/evaluations/export", headers=auth(org["hr"]),
                           params={"date_from": "2000-01-01"})
    wb2 = load_workbook(BytesIO(r_past.content))
    codes = [row[0].value for row in wb2["สรุป"].iter_rows(min_row=2)]
    assert "E1" in codes
